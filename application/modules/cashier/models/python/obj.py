import sys
import openpyxl
import os
import configparser
import sys
import json
import collections
import time
import base64
import pyodbc
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter

class Obj:

    custom_style = ""
    title_style = NamedStyle(name="title_style")
    title_style.font = Font(
        bold=True,
        size=11,
        color='2F4F4F'
    )
    title_style.alignment = Alignment(
        horizontal='center',
        vertical='center'
    )
    title_style.border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )

    title_style2 = NamedStyle(name="title_style2")
    title_style2.font = Font(
        bold=True,
        size=11,
        color='2F4F4F'
    )
    title_style2.alignment = Alignment(
        vertical='center'
    )
    title_style2.border = Border( 
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )

    def set_column_index(self, cursor, columns):

        i = 0
        columns_idx = []
        for column in cursor.description:
            for idx in columns:
                if (column[0] == idx):
                    columns[idx]["col"] = i 
            i = i + 1

        return columns_idx

    def set_header(self, ws, data, set_style = 0):

        current_row = ws._current_row + 1
        column = 1

        for h in data:
            column = 1
            if (len(data) > 1):
                for h2 in h:
                    current_pos = ws.cell(row=current_row, column=column, value=h2)

                    if (set_style == 1):
                        current_pos.style = self.title_style2
                    else:
                        current_pos.font = Font(bold=True)

                    column = column + 1
            else:
                for h2 in h:
                    current_pos = ws.cell(row=current_row, column=column, value=h2)

                    if (set_style == 1):
                        current_pos.style = self.title_style   
                    else:
                        current_pos.font = Font(bold=True)         

                    column =  column + 1              

            current_row = current_row + 1

    def set_footer(self, ws, data, columns):

        current_row = ws._current_row + 1
        column = 1

        i = 0
        for idx in columns.values():
            current_pos = ws.cell(row=current_row, column=(i + 1)) 
            value = data[0][i]
            current_pos.value = value
            current_pos.number_format = "#,##0.00"
            if (idx["format"] is not None):
                current_pos.number_format = idx["format"]
            else:
                current_pos.value
                current_pos.number_format = "#,##0.00"
            i = i + 1     

    def set_content(self, ws, data, columns):
        
        current_row = ws._current_row + 1
        for x in range(1, (len(data) + 1)):
            col = 1

            for y in columns.values():
                idx_row = x - 1
                idx_col = y["col"]

                current_cell = ws.cell(row=current_row, column=col)

                if y["format"] is not None:
                    current_cell.number_format = y["format"]

                current_cell.value = data[idx_row][idx_col]
                col = col + 1

            current_row = current_row + 1

    def set_content2(self, ws, content, column_config):

        for c in content:
            sp_name = c['sp_name']
            sp_param = c['sp_param']  
            definedValue = c['definedValue']
            current_row = ws._current_row + 1

            if (sp_name == "" and sp_param == ""):
                i = 1
                for v in definedValue:
                    ws.cell(row=current_row, column=i, value=v)
                    i = i+1
            else:
                pathCurrent = os.path.abspath(__file__ + "/../../../../../../")

                config = configparser.ConfigParser()
                config.read(pathCurrent+'/application/modules/main/configs/main.ini')

                server = config['production']['resources.db.params.host'][1:-1]
                database = 'cashier'
                username = config['production']['resources.db.params.username'][1:-1] 
                password = config['production']['resources.db.params.password'][1:-1]
                cnxn = pyodbc.connect("Driver={SQL Server};"
                                    "Server="+server+";Port=1433;"
                                    "Database="+database+";"
                                    "UID="+username+";"
                                    "PWD="+password+";")

                cursor = cnxn.cursor()
                # print("EXEC " + sp_name + " " + sp_param)
                # sys.exit()
                # execute query 
                cursor.execute("EXEC " + sp_name + " " + sp_param)
                data = cursor.fetchall()
                columns = [column[0] for column in cursor.description]

                self.set_column_index(cursor, column_config)
                self.set_content(ws, data, column_config)


        

    def set_style(self, ws, columns):

        i = 1
        for col in columns.values():
            col_letter = get_column_letter(i)
                        
            # Set Width
            if col["width"] is not None:
                ws.column_dimensions[col_letter].width = col["width"]
            else:
                ws.column_dimensions[col_letter].width = 15

            # Set Align
            if col["align"] is not None:
                ws.column_dimensions[col_letter].alignment = Alignment(horizontal=col["align"])
            else:
                ws.column_dimensions[col_letter].alignment = Alignment(horizontal="right")

            i = i + 1

    def set_bold_last_row(self, ws, isBold):
        for cell in ws[ws._current_row]:
            if (isBold == "true"):
                cell.font = Font(bold=True)
            else:
                cell.font = Font(bold=False)

    def set_custom_condition(self, ws, columns, condition):
        cols = list(columns)
        k = 1
        for col in cols:
            for con in condition:
                for cond in con:
                    if (col == cond['field']):
                        j = 1
                        for row in ws.iter_rows():
                            value = ws.cell(row=j, column=k).value

                            if (cond['condition'] == '' and value == None):
                                for cell in ws[j]:
                                    if (cond['bold'] == True):
                                        cell.font = Font(bold=True)
                            else:
                                if (cond['operator'] == '==' and cond['condition'] != ''):
                                    if (value == cond['condition']):
                                        for cell in ws[j]:
                                            if (cond['bold'] == True):
                                                cell.font = Font(bold=True)
                                elif (cond['operator'] == '>' and value is not None):
                                    if (int(value) > int(cond['condition'])):
                                        for cell in ws[j]:
                                            if (cond['bold'] == True):
                                                cell.font = Font(bold=True)
                                elif (cond['operator'] == '>=' and value is not None):
                                    if (int(value) >= int(cond['condition'])):
                                        for cell in ws[j]:
                                            if (cond['bold'] == True):
                                                cell.font = Font(bold=True)
                                elif (cond['operator'] == '<' and value is not None):
                                    if (int(value) < int(cond['condition'])):
                                        for cell in ws[j]:
                                            if (cond['bold'] == True):
                                                cell.font = Font(bold=True)
                                elif (cond['operator'] == '<=' and value is not None):
                                    if (int(value) <= int(cond['condition'])):
                                        for cell in ws[j]:
                                            if (cond['bold'] == True):
                                                cell.font = Font(bold=True)
                                elif (cond['operator'] == '!='):
                                    if (value != cond['condition']):
                                        for cell in ws[j]:
                                            if (cond['bold'] == True):
                                                cell.font = Font(bold=True)

                            j = j + 1 
            k = k + 1

    def set_remove_columns(self, ws, columns, removecolumns):
        cols = list(columns)
        k = 1
        for col in cols:
            for rm in removecolumns:
                if (col == rm):
                    ws.delete_cols(k, len(removecolumns))
            k = k + 1

    def set_merge_style(self, ws, merge_cell):
        for merge in merge_cell:
            ws.merge_cells(merge['from'] + ":" + merge['to'])
            ws[merge['from']].alignment = Alignment(vertical=merge['vertical_align'], horizontal=merge['horizontal_align'])

    def set_custom_style(self, ws, config, columns):
        for conf in config:

            if (conf['col'] != ""):
                cols = list(columns)
                k = 1
                for col in cols:
                    j = 1
                    for row in ws.iter_rows():
                        cell_value = ws.cell(row=j, column=k).value
                        if (cell_value == conf['val']):

                            alpha_num_from = ord(conf['from'].lower()) - 96    
                            alpha_num_until = ord(conf['to'].lower()) - 96 

                            for i in range(alpha_num_from, (alpha_num_until + 1)):
                                current = ws.cell(row=j, column=i)
                                current.alignment = Alignment(horizontal=conf['align'])
                                if (conf['bold'] == "True"):
                                    current.font = Font(bold=True)
                                else:
                                    current.font = Font(bold=False)

                                if (conf['format'] != "None"):
                                    current.number_format = conf['format']

                                if (conf['border'] == "True"):

                                    bdconfnone = Side(border_style=None, color='000000')

                                    borderTop = bdconfnone
                                    borderBottom = bdconfnone
                                    borderLeft = bdconfnone
                                    borderRight = bdconfnone
                                    
                                    borderConfig = conf['borderconfig']
                                    bdconf = Side(style=borderConfig['style'], color=borderConfig['bordercolor'])
                                    if (borderConfig['allborders'] == "True"):
                                        current.border = Border(top=bdconf, right=bdconf, bottom=bdconf, left=bdconf)
                                    else:
                                        if (borderConfig['bordertop'] == "True"):
                                            borderTop = bdconf
                                        elif (borderConfig['borderbottom'] == "True"):
                                            borderBottom = bdconf
                                        elif (borderConfig['borderleft'] == "True"):
                                            borderLeft = bdconf
                                        elif (borderConfig['borderright'] == "True"):
                                            borderRight = bdconf

                                        current.border = Border(top=borderTop, bottom=borderBottom, left=borderLeft, right=borderRight)
                        j = j+1
                    k = k+1
            else:
                l_col1 = ['', '']
                l_col2 = ['', '']

                tmp_list1 = list(conf['from'])
                for x in tmp_list1:
                    if (x.isalpha() == True):
                        l_col1[0] = l_col1[0] + x
                    else:
                        l_col1[1] = l_col1[1] + x

                tmp_list2 = list(conf['to'])
                for x in tmp_list2:
                    if (x.isalpha() == True):
                        l_col2[0] = l_col2[0] + x
                    else:
                        l_col2[1] = l_col2[1] + x

                alpha_num_from = ord(l_col1[0].lower()) - 96    
                alpha_num_until = ord(l_col2[0].lower()) - 96    

                for col_idx in range(alpha_num_from, (alpha_num_until + 1)):
                    for row_idx in range(int(l_col1[1]), (int(l_col2[1]) + 1)):
                        alpha = chr((col_idx + 96)).upper()
                        
                        current = ws[alpha + str(row_idx)]
                        current.alignment = Alignment(horizontal=conf['align'])
                        if (conf['bold'] == "True"):
                            current.font = Font(bold=True)
                        else:
                            current.font = Font(bold=False)

                        if (conf['format'] != "None"):
                            current.number_format = conf['format']

                        if (conf['border'] == "True"):

                            bdconfnone = Side(border_style=None, color='000000')

                            borderTop = bdconfnone
                            borderBottom = bdconfnone
                            borderLeft = bdconfnone
                            borderRight = bdconfnone
                            
                            borderConfig = conf['borderconfig']
                            bdconf = Side(style=borderConfig['style'], color=borderConfig['bordercolor'])
                            if (borderConfig['allborders'] == "True"):
                                current.border = Border(top=bdconf, right=bdconf, bottom=bdconf, left=bdconf)
                            else:
                                if (borderConfig['bordertop'] == "True"):
                                    borderTop = bdconf
                                elif (borderConfig['borderbottom'] == "True"):
                                    borderBottom = bdconf
                                elif (borderConfig['borderleft'] == "True"):
                                    borderLeft = bdconf
                                elif (borderConfig['borderright'] == "True"):
                                    borderRight = bdconf

                                current.border = Border(top=borderTop, bottom=borderBottom, left=borderLeft, right=borderRight)
                            

                

