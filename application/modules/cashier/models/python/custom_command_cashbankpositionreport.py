import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
import sys
import os

filename = sys.argv[1]
report_type = sys.argv[2]

path = os.path.abspath(__file__ + "/../../../../../../")
path = path+'/public/app/gl/uploads/' + filename

wb_obj = openpyxl.load_workbook(path)
for sheet in wb_obj.worksheets: 
    max_row = sheet._current_row
    
    if (sheet.title == 'Summary' and report_type == 'K'):
        sheet.delete_rows(21, 2)

    if (sheet.title == 'Cash Detail' or sheet.title == 'CashBank Detail'):
        sheet.insert_rows(idx=max_row, amount=4)
        for colidx in range(sheet.min_column, (sheet.max_column + 1)):
            column_name = sheet.cell(row=6, column=colidx).value
            if (column_name == 'SALDO AWAL' or column_name == 'DEBET' or column_name == 'KREDIT' or column_name == 'SALDO AKHIR'):
                c = sheet.cell(row=((max_row + 4) - 2), column=colidx)
                if (column_name == 'SALDO AWAL'):
                    c.value = 'Cash Ending Balance'
                elif (column_name == 'DEBET'):
                    c.value = 'Total Cash Advance'
                elif (column_name == 'KREDIT'):
                    c.value = 'Total Project Loan'
                elif (column_name == 'SALDO AKHIR'):
                    c.value = 'Total Cash Ending Balance'
                c.font = Font(bold=True)
                c.alignment = Alignment(horizontal='center')
                c.border = Border(bottom=Side(border_style='thick'))

wb_obj.save(path)