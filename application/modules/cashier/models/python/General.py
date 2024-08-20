import openpyxl
import os
import configparser
import sys
import json
import collections
import time
import base64
import pyodbc
from obj import Obj
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter

pathCurrent = os.path.abspath(__file__ + "/../../../../../../")

config = configparser.ConfigParser()
config.read(pathCurrent+'/application/modules/main/configs/main.ini')

server = config['production']['resources.db.params.host'][1:-1]
database = 'cashier'
username = config['production']['resources.db.params.username'][1:-1] 
password = config['production']['resources.db.params.password'][1:-1]
cnxn = pyodbc.connect("Driver={SQL Server};"
                      "Server="+server+";Port=1433;"
                      "Database="+database+";"
                      "UID="+username+";"
                      "PWD="+password+";")

cursor = cnxn.cursor()

base64_file = sys.argv[1]
filename = sys.argv[2]

path = os.path.abspath(__file__ + "/../../../../../../")
newFilePath = path+'/public/app/gl/uploads/' + filename

f = open(path+'/public/app/gl/uploads/' + base64_file, 'r')
parameter = f.read()

obj = json.loads(base64.b64decode(parameter))

wb = Workbook()
h = Obj()

for o in obj:

    content = o['content']

    # create sheet
    if (o['sheet'] == 0):
        ws = wb.active
        ws.title = content['title']

    else:
        ws = wb.create_sheet(content['title'])

    # set header or title
    h.set_header(ws, content['header'])

    if (len(content['data']['column_title']) > 0):
        h.set_header(ws, content['data']['column_title'], 1)    
    
    if (len(content['content']) > 0):
        h.set_content2(ws, content['content'], content['data']['column_config'])
    else:

        # execute query 
        cursor.execute("EXEC " + content['data']['sp'] + " " + content['data']['sp_param'])
        data = cursor.fetchall()
        columns = [column[0] for column in cursor.description]

        h.set_column_index(cursor, content['data']['column_config'])
        h.set_content(ws, data, content['data']['column_config'])
    
    # set footer
    if (len(content['footer']) > 0):
        if (content['footer_as_data'] == True):
            h.set_footer(ws, content['footer'], content['data']['column_config'])
        else:
            h.set_footer(ws, content['footer'], None)
    
    # set styling in data
    h.set_style(ws, content['data']['column_config'])

    # set merge style
    h.set_merge_style(ws, o['merge_cell'])

    # set bold in last active row
    h.set_bold_last_row(ws, o['style']['bold_last_row'])

    # set custom condition
    h.set_custom_condition(ws, content['data']['column_config'], o['condition'])

    # set custom style (bold, align, format, border)
    h.set_custom_style(ws, o['custom_style'], content['data']['column_config'])

    # remove specific columns
    if (o['remove_columns'] != None): 
        if (len(o['remove_columns']) > 0):
            h.set_remove_columns(ws, content['data']['column_config'], o['remove_columns'])

cnxn.close()

wb.save(newFilePath)
print(filename)